import openpyxl
from openpyxl.styles import PatternFill
from tkinter import messagebox

# 파일 열기
wb = openpyxl.load_workbook('your_file.xlsx')
sheet = wb.active

# 바꿀 단어 설정
replacements = {
    '1) 향선택: 메르헨트 디퓨저 500ml 2개 런드리룸': '♡4W_OMHA1026_섬유향수 오드 만다린 250ml 2개입',

    # 이런 식으로 계속 추가해나가세요.
}

# 파란색 스타일 설정
blue_fill = PatternFill(start_color="0000FF",
                        end_color="0000FF", fill_type="solid")

# N열 2행부터 값 비교 및 변경
# N열은 엑셀에서 14번째 열입니다.

# N열 2행부터 3000행까지 값 비교 및 변경
# N열은 엑셀에서 14번째 열입니다.
for idx, row in enumerate(sheet.iter_rows(min_row=2, min_col=14, values_only=True), start=2):
    if row[0] in replacements:
        sheet.cell(row=idx, column=14, value=replacements[row[0]])
    else:
        sheet.cell(row=idx, column=14).fill = blue_fill

# 변경된 내용 저장
wb.save('output_file.xlsx')
messagebox.showinfo("완료", "작업이 완료되었습니다!")
