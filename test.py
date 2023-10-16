import openpyxl

replacements = {
    '1) 향선택: 메르헨트 디퓨저 500ml 2개 런드리룸 ': '♡4W_OMHA1026_섬유향수 오드 만다린 250ml 2개입',


}

wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.active

# for idx, row in enumerate(rows[1:], start=2):  # 첫 번째 행을 건너뜁니다.
#     if row[13].value in replacements:  # 열 'N'의 인덱스가 13인 경우, 0부터 시작함에 유의하세요.
#         row[13].value = replacements[row[13].value]
#     else:
#         for cell in rows[idx]:
#             cell.fill = openpyxl.styles.PatternFill(
#                 start_color="FF0000", end_color="FF0000", fill_type="solid")

# for idx, row in enumerate(rows):
#     if idx == 0:  # 첫 번째 행은 헤더입니다.
#         continue
#     if row[13].value in replacements:
#         row[13].value = replacements[row[13].value]
#     else:
#         for cell in row:
#             cell.fill = openpyxl.styles.PatternFill(
#                 start_color="FF0000", end_color="FF0000", fill_type="solid")

for row in sheet.iter_rows(min_row=2, values_only=True):
    if row[13] in replacements:
        row[13] = replacements[row[13]]
    else:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(
                start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")


wb.save('output_file.xlsx')
